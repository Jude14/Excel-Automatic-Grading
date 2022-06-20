from openpyxl import Workbook, workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font #to style things

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title="Grades"

headings =  ['Name']+ list(data['Joe'].keys())+['std grd']
ws.append(headings)

print()
for person in data:
    grades = list(data[person].values())
    ws.append([person]+ grades)
ws.append(['subject avg'])

for col in range(2,len(data['Joe'])+2):
    char = get_column_letter(col)
    ws[char+str(len(data)+2)]= f"=SUM({char+'2'}:{char+str(len(data['Joe'])+2)})/{len(data)}"

for row in range(2,len(data)+2):
    ws[get_column_letter(len(data['Joe'])+2)+str(row)]=f"=SUM({'B'+str(row)}:{'E'+str(row)})/{len(data['Joe'])}"

for col in range(1, len(data)+2):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("NewGrades.xlsx") 












'''
#baamol wb = load_workbook('esm_l_file.xlsx')  eza badde eftah file excel mawjoud
wb = load_workbook('T1.xlsx')
ws = wb.active
ws.merge_cells("A1:D2") #merge all the cells in this range
ws.unmerge_cells("A1:D2")
ws.insert_rows(6) #insert empty row at row 6
ws.delete_rows(6) #del row 6
ws.insert_cols(2) #insert empty columns at column B
ws.delete_cols(2)
ws.move_range("A1:D4",rows = 4,cols = 2) #shift whats in the specified range
wb.save('T1.xlsx')'''

'''for row in range (1,11):
    for col in range (1,5):
        char = get_column_letter(col)
        ws[char+ str(row)] = char + str(row) #loop through the rows and columns
        wb.save('Grades_Trial.xlsx')'''

'''wb = Workbook()  #to create and work on a new excel file
ws = wb.active
ws.title = 'Data'
ws.append(['just','trying','this','out'])   #adding a python list as a row
wb.save("T1.xlsx") #save as new excel file named T1'''

'''wb = load_workbook('Grades_Trial.xlsx')
wb.create_sheet("sheetXXX")
print(wb.sheetnames)'''
'''ws = wb.active
print(ws['A1'].value)
print(wb.sheetnames)
'''
'''ws['A1'].value = 'test'
ws['E2']=  ws['A2'].value
wb.save('Grades_Trial.xlsx')'''